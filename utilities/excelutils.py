import openpyxl

# 엑셀 파일의 행수를 계산해서 반환해줌 
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

# 엑셀 파일의 열 수를 계산해서 반환해줌 
def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

# 엑셀 파일의 행번호, 열번호에 있는 데이터 값을 반환해줌
def readData(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=columnNum).value

# data 값을 행번호, 열번호에 쓰고 파일을 저장함 
def writeData(file, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=columnNum).value = data
    workbook.save(file)